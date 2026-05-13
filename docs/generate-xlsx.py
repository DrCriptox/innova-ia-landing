"""
Genera docs/requisitos-calificacion.xlsx — tabla §7 editable.
v3: Estructura actualizada con
- 12 rangos nuevos (Consultor → Leyenda) con nombres y PVs revisados
- Columna "PV/mes 6 niveles" (puntos de volumen, 1 PV = $2 USD)
- Columna "Bono Gift" editable (default 0)
- Columna "Pool 5%" (acciones del pool de rangos 5%)
- Columna "Pool Aprox $" (monto estimado en USD)
- Columna "Cobro estimado" como FÓRMULA:
  min = 22% × PV current + Bono Gift + Pool Aprox $
  max = 22% × PV next    + Bono Gift + Pool Aprox $
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Requisitos de Calificación"

# Colores y estilos
GOLD = "B8841F"
DARK = "1A1A1A"
WHITE = "FFFFFF"
BORDER = "D4D4D4"

thin = Side(border_style="thin", color=BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
header_fill = PatternFill("solid", fgColor=DARK)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_font = Font(name="Calibri", size=11)
cell_align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

rank_font = Font(name="Calibri", size=11, bold=True)
gold_font = Font(name="Calibri", size=11, bold=True, color=GOLD)
green_font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
elite_fill = PatternFill("solid", fgColor="FFF8E8")
leyenda_fill = PatternFill("solid", fgColor="FCE4A6")

input_fill = PatternFill("solid", fgColor="EAF4FF")
input_fill_elite = PatternFill("solid", fgColor="FFF3D6")
input_fill_leyenda = PatternFill("solid", fgColor="FCDD8E")

# Título
ws.merge_cells("A1:J1")
ws["A1"] = "§7 REQUISITOS DE CALIFICACIÓN — Plan de Carrera Sky"
ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=DARK)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:J2")
ws["A2"] = "Cada rango exige cumplir TODOS los criterios en el mismo mes calendario · PROPUESTA — REVISAR"
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

ws.row_dimensions[3].height = 8

# Headers
headers = ["#", "Rango", "PV/mes 6 niveles", "PV mínimo", "Clientes activos",
           "Rangos directos", "Bono Gift", "Pool 5%", "Pool Aprox $", "Cobro estimado"]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
ws.row_dimensions[4].height = 36

# Datos: rango, PV/mes, PV min, clientes, rangos directos, bono gift, pool shares, pool aprox $
data = [
    (1,  "Consultor",              750,    200,   1, 0, 0,    0.2,   0),
    (2,  "Ejecutivo",             1500,    200,   2, 0, 150,  0.5,   24),
    (3,  "Director",              4000,    300,   2, 1, 400,  2,     96),
    (4,  "Director Ejecutivo",    8000,    300,   3, 2, 550,  4,     192),
    (5,  "Director Global",      15000,    400,   3, 3, 800,  8,     384),
    (6,  "Presidente",           35000,    500,   4, 3, 1600, 15,    720),
    (7,  "Presidente Ejecutivo", 70000,    500,   4, 4, 2800, 25,    1200),
    (8,  "Presidente Global",   150000,    500,   5, 4, 3800, 50,    2400),
    (9,  "Embajador",           300000,   1000,   5, 5, 4000, 100,   4800),
    (10, "Embajador Global",    500000,   1000,   6, 5, 4000, 150,   7200),
    (11, "Icon",               1000000,   1000,   6, 6, 4000, 250,   12000),
    (12, "Leyenda",            2000000,   1000,   7, 6, 4000, 400,   19200),
]

start_row = 5
last_idx = len(data) - 1

for i, row in enumerate(data):
    r = start_row + i
    num, rango, pv_mes, pv_min, clientes, rangos_dir, bono, pool_shares, pool_aprox = row

    # Numero
    c = ws.cell(row=r, column=1, value=num)
    c.font = cell_font
    c.alignment = cell_align_center

    # Rango
    c = ws.cell(row=r, column=2, value=rango)
    c.font = rank_font
    c.alignment = cell_align_left

    # PV/mes 6 niveles
    c = ws.cell(row=r, column=3, value=pv_mes)
    c.font = gold_font
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '#,##0" PV"'

    # PV mínimo
    c = ws.cell(row=r, column=4, value=pv_min)
    c.font = cell_font
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '"$"#,##0'

    # Clientes activos
    c = ws.cell(row=r, column=5, value=clientes)
    c.font = cell_font
    c.alignment = cell_align_center

    # Rangos directos
    c = ws.cell(row=r, column=6, value=rangos_dir)
    c.font = cell_font
    c.alignment = cell_align_center

    # Bono Gift (editable)
    c = ws.cell(row=r, column=7, value=bono)
    c.font = Font(name="Calibri", size=11, color="0066CC", bold=True)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '"$"#,##0'

    # Pool 5% (shares como número con 'ac')
    c = ws.cell(row=r, column=8, value=pool_shares)
    c.font = cell_font
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '0.0"ac";"0ac"'

    # Pool Aprox $
    c = ws.cell(row=r, column=9, value=pool_aprox)
    c.font = cell_font
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.number_format = '"$"#,##0'

    # Cobro estimado (FÓRMULA: 22% × PV/mes + Bono Gift + Pool Aprox $)
    if i < last_idx:
        formula = (
            f'=TEXT(INT(0.22*C{r})+G{r}+I{r},"$#,##0")&" – "&'
            f'TEXT(INT(0.22*C{r+1})+G{r}+I{r},"$#,##0")'
        )
    else:
        formula = f'=TEXT(INT(0.22*C{r})+G{r}+I{r},"$#,##0")&"+"'
    c = ws.cell(row=r, column=10, value=formula)
    c.font = green_font
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Bordes
    for col in range(1, 11):
        ws.cell(row=r, column=col).border = border

    # Fills por tier
    if num in (9, 10, 11):
        for col in range(1, 11):
            if col != 7:
                ws.cell(row=r, column=col).fill = elite_fill
        ws.cell(row=r, column=7).fill = input_fill_elite
    elif num == 12:
        for col in range(1, 11):
            if col != 7:
                ws.cell(row=r, column=col).fill = leyenda_fill
        ws.cell(row=r, column=7).fill = input_fill_leyenda
    else:
        ws.cell(row=r, column=7).fill = input_fill

    ws.row_dimensions[r].height = 26

# Definiciones
defs_row = start_row + len(data) + 2
ws.cell(row=defs_row, column=1, value="DEFINICIONES").font = Font(bold=True, size=11, color=GOLD)
defs_row += 1

definitions = [
    ("PV (Punto de Volumen)",     "Unidad de medida del plan. Equivalencia oficial: 1 PV = $2 USD."),
    ("PV/mes 6 niveles",          "Volumen total facturado por toda tu estructura de 6 niveles uninivel en el mes."),
    ("PV mínimo",                 "Volumen Personal mínimo que TÚ debes producir cada mes (USD)."),
    ("Clientes activos",          "Socios que TÚ trajiste personalmente y cumplen su PV mínimo ese mes."),
    ("Rangos directos",           "Patrocinados directos que mantienen rango Director o superior."),
    ("Bono Gift (editable)",      "Bono adicional discrecional. Editas en USD (celda azul) y se suma al Cobro."),
    ("Pool 5% (shares)",          "Acciones del Pool de Rangos 5%. Más rango = más shares = más cobro del pool."),
    ("Pool Aprox $",              "Estimación del cobro mensual del Pool 5% según shares y volumen global."),
    ("Cobro estimado (fórmula)",  "Min = 22% × PV/mes (rango actual) + Bono Gift + Pool Aprox $. Max = 22% × PV/mes (rango siguiente) + Bono Gift + Pool Aprox $. Leyenda solo muestra el mínimo con '+'."),
    ("Requisito de Balance",      "Para mantener uninivel completo (N4-N6), tu pierna mayor no puede aportar más del 75% del PV/mes. Aplica desde Director."),
]

for term, desc in definitions:
    c = ws.cell(row=defs_row, column=1, value=term)
    c.font = Font(bold=True, size=10)
    c.alignment = Alignment(vertical="top")
    ws.merge_cells(start_row=defs_row, start_column=2, end_row=defs_row, end_column=10)
    c = ws.cell(row=defs_row, column=2, value=desc)
    c.font = Font(size=10, color="444444")
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[defs_row].height = 24
    defs_row += 1

# Pendientes
defs_row += 1
ws.cell(row=defs_row, column=1, value="PENDIENTES DE CONFIRMAR").font = Font(bold=True, size=11, color=GOLD)
defs_row += 1

pending = [
    "PV mínimo por rango (¿OK los valores propuestos?)",
    "Bono Gift por rango (¿confirmamos los montos?)",
    "Pool 5% shares por rango (escala 1.5x-2x — ¿ajustar?)",
    "Pool Aprox $ — depende del volumen global mensual; aquí se muestra estimado",
    "Activar/no activar Requisito de Balance 75% desde Director",
]

for p in pending:
    ws.merge_cells(start_row=defs_row, start_column=1, end_row=defs_row, end_column=10)
    c = ws.cell(row=defs_row, column=1, value="☐  " + p)
    c.font = Font(size=10)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[defs_row].height = 20
    defs_row += 1

# Anchos
widths = [5, 22, 16, 12, 12, 12, 12, 10, 12, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

# Pie de página
ws.oddFooter.center.text = "Innova IA · Sky System — Plan de Carrera · v3 — Mayo 2026"
ws.oddFooter.center.size = 9
ws.oddFooter.center.color = "888888"

output = "/home/user/innova-ia-landing/docs/requisitos-calificacion.xlsx"
wb.save(output)
print(f"Generado: {output}")
